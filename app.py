from __future__ import annotations

import base64
import csv
import io
import json
import os
import random
import re
import ssl
import threading
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from flask import Flask, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "enriquecimento-demo-secret-key"

@app.route('/favicon.ico')
@app.route('/favicon.png')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.png', mimetype='image/png')

USERS = {
    "admin": {"password": "admin123", "role": "admin", "name": "Estratégia - Admin"},
    "saf": {"password": "saf123", "role": "user1", "name": "Usuário SAF"},
}

BASE_DIR = Path(__file__).resolve().parent
import tempfile

# On Vercel/Serverless, the project folder is usually read-only.
# Keep runtime files and exports under /tmp to avoid FUNCTION_INVOCATION_FAILED.
is_vercel = os.environ.get("VERCEL") == "1" or os.environ.get("VERCEL_ENV") is not None
TEMP_PARENT = Path(tempfile.gettempdir()) if is_vercel else BASE_DIR

UPLOAD_DIR = TEMP_PARENT / "uploads"
EXPORT_DIR = TEMP_PARENT / "exports"
LOG_PATH = TEMP_PARENT / "api_debug.log"
UPLOAD_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

STATE: Dict[str, List[Dict[str, Any]]] = {
    "records": [],
    "queue": [],
    "notifications": [],
}

SUPABASE_URL = "https://yaresjmqcrpuiorpvbck.supabase.co/rest/v1/"
SUPABASE_KEY = "sb_publishable_k7ISy3EG7LO6aR7Htsx7_g_1DIWf7OP"

def make_supabase_request(table: str, method: str = "GET", data: Any = None, query: str = "") -> Any:
    url = f"{SUPABASE_URL}{table}{query}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    
    encoded_data = None
    if data is not None:
        encoded_data = json.dumps(data).encode("utf-8")
        
    req = urllib.request.Request(
        url,
        data=encoded_data,
        headers=headers,
        method=method
    )
    
    try:
        context = ssl._create_unverified_context()
        with urllib.request.urlopen(req, timeout=10, context=context) as resp:
            body = resp.read().decode("utf-8")
            if body:
                return json.loads(body)
            return []
    except Exception as e:
        log_api_debug(f"Supabase request failed ({method} {url}): {e}")
        return None

def db_save_record(record: Dict[str, Any]) -> None:
    record_db = {
        k: v for k, v in record.items()
        if k in {"id", "filename", "stored_file", "type", "available_fields", "preview", "total_in", "total_enriched", "columns", "invalid_phone_count", "invalid_phones"}
    }
    res = make_supabase_request("records", "POST", record_db)
    if res is None:
        log_api_debug("Fallback to in-memory STATE['records']")
    # Always keep in memory state as well for compatibility / fallback
    STATE["records"].append(record)

def db_get_record(record_id: str) -> Dict[str, Any] | None:
    res = make_supabase_request("records", "GET", query=f"?id=eq.{record_id}")
    if res and len(res) > 0:
        return res[0]
    return next((r for r in STATE["records"] if r["id"] == record_id), None)

def extract_reject_reason(item: Dict[str, Any]) -> str:
    if item.get("reject_reason"):
        return str(item["reject_reason"]).strip()
    logs = item.get("logs") or []
    for log in logs:
        if isinstance(log, str) and "Solicitação recusada pela Estratégia. Motivo: " in log:
            return log.split("Solicitação recusada pela Estratégia. Motivo: ")[1].strip()
    return ""

def db_save_queue_item(item: Dict[str, Any]) -> None:
    item_db = {
        k: v for k, v in item.items()
        if k in {"id", "record_id", "queue_number", "requester_name", "observacoes", "filename", "status", "total_rows", "processed_count", "success_count", "error_count", "request_time", "completed_at", "summary_name", "selected_fields", "detected_type", "logs"}
    }
    res = make_supabase_request("queue", "POST", item_db)
    if res is None:
        log_api_debug("Fallback to in-memory STATE['queue']")
    STATE["queue"].append(item)

def db_update_queue_item(item_id: str, updates: Dict[str, Any]) -> None:
    updates_db = {
        k: v for k, v in updates.items()
        if k in {"id", "record_id", "queue_number", "requester_name", "observacoes", "filename", "status", "total_rows", "processed_count", "success_count", "error_count", "request_time", "completed_at", "summary_name", "selected_fields", "detected_type", "logs"}
    }
    res = make_supabase_request("queue", "PATCH", data=updates_db, query=f"?id=eq.{item_id}")
    if res is None:
        log_api_debug("Fallback to update in-memory STATE['queue']")
    # Update local in-memory
    item = next((job for job in STATE["queue"] if job["id"] == item_id), None)
    if item:
        item.update(updates)

def db_get_queue_item(item_id: str) -> Dict[str, Any] | None:
    res = make_supabase_request("queue", "GET", query=f"?id=eq.{item_id}")
    item = None
    if res and len(res) > 0:
        item = res[0]
    else:
        item = next((job for job in STATE["queue"] if job["id"] == item_id), None)
    if item:
        item["reject_reason"] = extract_reject_reason(item)
    return item

def db_get_all_queue() -> List[Dict[str, Any]]:
    res = make_supabase_request("queue", "GET")
    items = res if res is not None else STATE["queue"]
    for item in items:
        item["reject_reason"] = extract_reject_reason(item)
    if res is not None:
        STATE["queue"] = res
    return items

def db_save_notification(notification: Dict[str, Any]) -> None:
    res = make_supabase_request("notifications", "POST", notification)
    if res is None:
        log_api_debug("Fallback to in-memory STATE['notifications']")
    STATE["notifications"].append(notification)

def db_get_notifications() -> List[Dict[str, Any]]:
    res = make_supabase_request("notifications", "GET")
    if res is not None:
        STATE["notifications"] = res
        return res
    return STATE["notifications"]

REQUIRED_FIELDS = ["nome", "telefone"]
EXTRA_FIELDS = [
    "cpf",
    "data_nascimento",
    "idade",
    "sexo",
    "email",
    "logradouro",
    "numero",
    "complemento",
    "bairro",
    "cidade",
    "estado",
    "cep",
    "renda_estimada",
    "cbo",
    "profissao",
    "score",
    "perfil",
    "obito",
    "whatsapp",
    "status"
]
ALL_ENRICH_FIELDS = REQUIRED_FIELDS + EXTRA_FIELDS


def load_local_config() -> Dict[str, str]:
    config_file = Path(__file__).resolve().parent / "config.json"
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    env_file = Path(__file__).resolve().parent / ".env"
    if env_file.exists():
        config = {}
        try:
            with open(env_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        config[k.strip()] = v.strip().strip('"').strip("'")
            return config
        except Exception:
            pass

    return {}

LOCAL_CONFIG = load_local_config()

NOVA_VIDA_BASE_URL = (LOCAL_CONFIG.get("NOVA_VIDA_BASE_URL") or os.getenv("NOVA_VIDA_BASE_URL") or "https://wsnv.novavidati.com.br").strip().rstrip("/")
NOVA_VIDA_TOKEN_URL = (LOCAL_CONFIG.get("NOVA_VIDA_TOKEN_URL") or os.getenv("NOVA_VIDA_TOKEN_URL") or f"{NOVA_VIDA_BASE_URL}/WSLocalizador.asmx/GerarToken").strip()
NOVA_VIDA_API_URL = (LOCAL_CONFIG.get("NOVA_VIDA_API_URL") or os.getenv("NOVA_VIDA_API_URL") or NOVA_VIDA_BASE_URL).strip().rstrip("/")
NOVA_VIDA_API_KEY = (LOCAL_CONFIG.get("NOVA_VIDA_API_KEY") or os.getenv("NOVA_VIDA_API_KEY") or os.getenv("NOVA_VIDA_TOKEN") or "").strip()
NOVA_VIDA_API_ENDPOINT = (LOCAL_CONFIG.get("NOVA_VIDA_API_ENDPOINT") or os.getenv("NOVA_VIDA_API_ENDPOINT") or "WSLocalizador.asmx/PesquisaAtributosUnico").strip().lstrip("/")
NOVA_VIDA_USER = (LOCAL_CONFIG.get("NOVA_VIDA_USER") or os.getenv("NOVA_VIDA_USER") or "").strip()
NOVA_VIDA_PASSWORD = (LOCAL_CONFIG.get("NOVA_VIDA_PASSWORD") or os.getenv("NOVA_VIDA_PASSWORD") or "").strip()
NOVA_VIDA_CLIENT = (LOCAL_CONFIG.get("NOVA_VIDA_CLIENT") or os.getenv("NOVA_VIDA_CLIENT") or "").strip()
NOVA_VIDA_API_HEADER = (LOCAL_CONFIG.get("NOVA_VIDA_API_HEADER") or os.getenv("NOVA_VIDA_API_HEADER") or "Authorization").strip()
NOVA_VIDA_API_METHOD = (LOCAL_CONFIG.get("NOVA_VIDA_API_METHOD") or os.getenv("NOVA_VIDA_API_METHOD") or "POST").strip().upper()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace(" ", "_")
    text = re.sub(r"[^a-z0-9_]+", "_", text)
    return text.strip("_")


def labelize(value: str) -> str:
    safe = (value or "").strip()
    if not safe:
        return "Campo"
    if safe.lower() == "cpf":
        return "CPF"
    if safe.lower() == "cbo":
        return "CBO"
    if safe.lower() in ("uf", "estado"):
        return "Estado"
    safe = safe.replace("_", " ")
    return safe.title()


def format_to_brazilian_date(val: Any) -> str:
    if val is None or val == "":
        return ""
    val_str = str(val).strip()
    # DD/MM/YYYY
    if re.match(r"^\d{2}/\d{2}/\d{4}$", val_str):
        return val_str
    # YYYY-MM-DD
    match_ymd = re.match(r"^(\d{4})[-/](\d{2})[-/](\d{2})$", val_str)
    if match_ymd:
        y, m, d = match_ymd.groups()
        return f"{d}/{m}/{y}"
    # DD-MM-YYYY
    match_dmy = re.match(r"^(\d{2})[-/](\d{2})[-/](\d{4})$", val_str)
    if match_dmy:
        d, m, y = match_dmy.groups()
        return f"{d}/{m}/{y}"
    # YYYYMMDD
    match_compact = re.match(r"^(\d{4})(\d{2})(\d{2})$", val_str)
    if match_compact:
        y, m, d = match_compact.groups()
        return f"{d}/{m}/{y}"
    return val_str


def normalize_email_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_renda_estimada(value: Any) -> str:
    if value is None or value == "":
        return ""
    text = str(value).strip()
    if not text:
        return ""

    if re.search(r"\d[\d\.,]*\s*(?:AT|A T)\s*\d[\d\.,]*", text, flags=re.IGNORECASE):
        return re.sub(r"\s+", " ", text).replace("  ", " ")

    if text.upper().startswith("R$") or text.upper().startswith("RS"):
        text = text[2:] if text.upper().startswith("R$") else text[2:]
    cleaned = re.sub(r"[^\d,\.\-\s]", "", text)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return text
    return cleaned


def format_concatenated_fields(row: Dict[str, Any], selected_fields: List[str]) -> str:
    nome_val = row.get("nome") or ""
    tel_val = row.get("telefone") or ""
    
    result = f"{nome_val}|{tel_val}|"
    
    remaining = [f for f in selected_fields if f not in ("nome", "telefone")]
    if remaining:
        parts = []
        for field in remaining:
            val = row.get(field)
            if val is None or val == "":
                val = ""
            parts.append(f"{labelize(field)}: {val}")
        result += " " + " ".join(parts)
    return result


def detect_enrichment_type(columns: List[str]) -> str | None:
    combined = " ".join(normalize_text(col) for col in columns)
    has_nome = "nome" in combined
    has_uf = "uf" in combined or "estado" in combined
    has_cpf = "cpf" in combined
    has_telefone = any(t in combined for t in ["telefone", "celular", "fone", "tel"])
    
    if has_cpf:
        return "CPF"
    if has_nome and has_uf:
        return "NOME_UF"
    if has_nome and has_telefone:
        return "NOME_TELEFONE"
    if has_nome:
        return "NOME"
    if has_telefone:
        return "TELEFONE"
    if has_uf:
        return "UF"
    return None


def generate_mock_row(row: Dict[str, Any], detected_type: str) -> Dict[str, Any]:
    name = row.get("nome") or row.get("nome_completo") or row.get("nome_cliente") or f"Contato {uuid.uuid4().hex[:6]}"
    
    # Gerar DDD e telefone aleatórios reais se não informados
    ddd = random.choice(["11", "12", "19", "21", "22", "27", "31", "32", "41", "47", "51", "61", "62", "71", "81", "85", "91", "98"])
    raw_phone = row.get("telefone") or row.get("celular") or row.get("telefone_contato") or f"{ddd}9{random.randint(6000, 9999)}{random.randint(1000, 9999)}"
    phone = re.sub(r"\D", "", str(raw_phone))
    
    cpf_value = row.get("cpf") or f"{(uuid.uuid4().int % 99999999999):011d}"
    
    # Lista de CBO e Profissões associadas
    cbo_professions = [
        ("411005", "Auxiliar de Escritório"),
        ("782310", "Motorista de Furgão ou Veículo Similar"),
        ("225250", "Médico Ginecologista e Obstetra"),
        ("322205", "Técnico de Enfermagem"),
        ("142105", "Gerente Administrativo"),
        ("212405", "Analista de Desenvolvimento de Sistemas"),
        ("231205", "Professor da Educação Infantil"),
        ("354125", "Supervisor de Vendas Comercial")
    ]
    cbo_pair = random.choice(cbo_professions)
    cbo_value = row.get("cbo") or cbo_pair[0]
    profession_value = row.get("profissao") or row.get("possivel_profissao") or cbo_pair[1]
    
    # Gerar idade e data de nascimento aleatórias realistas
    age = random.randint(18, 65)
    year = datetime.now().year - age
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    data_nascimento = format_to_brazilian_date(row.get("data_nascimento") or f"{day:02d}/{month:02d}/{year}")
    
    sex_value = row.get("sexo") or random.choice(["MASCULINO", "FEMININO"])
    
    clean_name_part = normalize_text(name).replace("_", "")
    email_val = row.get("email") or f"{clean_name_part[:12]}@{random.choice(['gmail.com', 'hotmail.com', 'outlook.com', 'yahoo.com.br'])}"
    email_val = normalize_email_value(email_val)
    
    # Endereços realistas
    streets = [
        "Luiz Rodrigues dos Santos", "Maria Matos", "Doutor Joao Mangabeira", 
        "Analia Ferreira Ferro", "Avenida Brasil", "Rua das Flores", 
        "Rua Rio de Janeiro", "Avenida Paulista"
    ]
    neighborhoods = [
        "Todos os Santos", "Centro", "Brasilia", "Praia do Morro", 
        "Jardim Paulista", "Copacabana", "Boa Viagem", "Savassi"
    ]
    
    logradouro_val = row.get("logradouro") or random.choice(streets)
    numero_val = row.get("numero") or str(random.randint(10, 1500))
    complemento_val = row.get("complemento") or random.choice(["AP 401", "AP 301", "Casa A", "Bloco B", "SL 12", ""])
    bairro_val = row.get("bairro") or random.choice(neighborhoods)
    cep_val = row.get("cep") or f"{random.randint(10000, 99999):05d}{random.randint(100, 999):03d}"
    
    # Cidades e estados coordenados do Brasil
    cities_states = [
        ("São Paulo", "SP"),
        ("Campinas", "SP"),
        ("Santos", "SP"),
        ("Rio de Janeiro", "RJ"),
        ("Niterói", "RJ"),
        ("Belo Horizonte", "MG"),
        ("Uberlândia", "MG"),
        ("Curitiba", "PR"),
        ("Londrina", "PR"),
        ("Porto Alegre", "RS"),
        ("Caxias do Sul", "RS"),
        ("Salvador", "BA"),
        ("Feira de Santana", "BA"),
        ("Recife", "PE"),
        ("Olinda", "PE"),
        ("Fortaleza", "CE"),
        ("Brasília", "DF"),
        ("Goiânia", "GO"),
        ("Manaus", "AM"),
        ("Belém", "PA"),
    ]
    
    city_val = row.get("cidade")
    state_val = row.get("uf") or row.get("estado")
    
    if not city_val and not state_val:
        city_val, state_val = random.choice(cities_states)
    elif state_val and not city_val:
        state_upper = str(state_val).strip().upper()
        matching = [c for c, s in cities_states if s == state_upper]
        city_val = random.choice(matching) if matching else f"Cidade de {state_val}"
    elif city_val and not state_val:
        city_lower = normalize_text(city_val).lower()
        matching = [s for c, s in cities_states if normalize_text(c).lower() == city_lower]
        state_val = random.choice(matching) if matching else "SP"
        
    renda_val = normalize_renda_estimada(
        row.get("renda_estimada")
        or row.get("renda")
        or random.choice([
            "1.000 AT 2.000",
            "2.001 AT 3.000",
            "3.001 AT 5.000",
            "5.001 AT 8.000",
            "18.001 AT 19.000",
        ])
    )
    
    score_val = row.get("score") or str(random.randint(200, 990))
    
    perfil_list = ["Sempre Presente", "O Bem Amado", "Quem Sou Eu", "Decisor", "Influenciador"]
    perfil = row.get("perfil") or row.get("personadecredito") or random.choice(perfil_list)
    
    status_list = ["ativo", "inativo", "pendente"]
    status = row.get("status") or random.choice(status_list)
    
    obito_val = row.get("obito") or row.get("flag_de_obito") or random.choices(["0", "1"], weights=[98, 2])[0]
    whatsapp_val = row.get("whatsapp") or random.choices(["SIM", "NÃO"], weights=[90, 10])[0]

    normalized = {
        "nome": name,
        "telefone": phone,
        "cpf": cpf_value,
        "data_nascimento": data_nascimento,
        "idade": str(age),
        "sexo": sex_value,
        "email": email_val,
        "logradouro": logradouro_val,
        "numero": numero_val,
        "complemento": complemento_val,
        "bairro": bairro_val,
        "cidade": city_val,
        "estado": state_val,
        "cep": cep_val,
        "renda_estimada": renda_val,
        "cbo": cbo_value,
        "profissao": profession_value,
        "score": score_val,
        "perfil": perfil,
        "obito": obito_val,
        "whatsapp": whatsapp_val,
        "status": status,
        "tipo_enriquecimento": detected_type,
    }

    for key, value in list(row.items()):
        normalized.setdefault(normalize_text(key), value)

    return normalized


def clean_dataframe(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    cleaned_rows: List[Dict[str, Any]] = []
    for row in rows:
        cleaned_row: Dict[str, Any] = {}
        for key, value in row.items():
            clean_name = normalize_text(key)
            if not clean_name:
                clean_name = f"campo_{len(cleaned_row) + 1}"
            cleaned_row[clean_name] = value

        # Merge DDD + Telefone/Celular columns if present
        ddd_key = next((k for k in cleaned_row.keys() if "ddd" in k), None)
        if ddd_key:
            ddd_val = re.sub(r"\D", "", str(cleaned_row[ddd_key])) if cleaned_row[ddd_key] else ""
            phone_keys = [k for k in cleaned_row.keys() if k != ddd_key and any(p in k for p in ["telefone", "celular", "fone", "cel"])]
            for p_key in phone_keys:
                if cleaned_row[p_key]:
                    phone_val = re.sub(r"\D", "", str(cleaned_row[p_key]))
                    if ddd_val and not phone_val.startswith(ddd_val):
                        cleaned_row[p_key] = f"{ddd_val}{phone_val}"
                    else:
                        cleaned_row[p_key] = phone_val
        else:
            # Clean existing phone digits anyway
            phone_keys = [k for k in cleaned_row.keys() if any(p in k for p in ["telefone", "celular", "fone", "cel"])]
            for p_key in phone_keys:
                if cleaned_row[p_key]:
                    cleaned_row[p_key] = re.sub(r"\D", "", str(cleaned_row[p_key]))

        # Filter out rows where the telephone is empty or only consists of zeros ("0", "000000", etc.)
        phone_keys = [k for k in cleaned_row.keys() if any(p in k for p in ["telefone", "celular", "fone", "cel"])]
        if phone_keys:
            has_valid_phone = False
            for p_key in phone_keys:
                val_digits = re.sub(r"\D", "", str(cleaned_row[p_key] or ""))
                if val_digits and not re.match(r"^0+$", val_digits):
                    has_valid_phone = True
                    break
            if not has_valid_phone:
                continue

        cleaned_rows.append(cleaned_row)
    return cleaned_rows


def normalize_name_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def build_enriched_rows(rows: List[Dict[str, Any]], detected_type: str) -> List[Dict[str, Any]]:
    enriched_rows: List[Dict[str, Any]] = []
    for row in rows:
        enriched = generate_mock_row(row, detected_type)
        enriched_rows.append(enriched)
    return enriched_rows


def first_present_value(record: Dict[str, Any], possible_keys: List[str]) -> Any:
    for key in possible_keys:
        value = record.get(key)
        if value is not None and value != "":
            return value
    return None


def normalize_api_row(row: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {}
    aliases = {
        "nome": ["nome", "nome_completo", "nome_cliente", "cliente", "name"],
        "telefone": ["telefone", "celular", "telefone_contato", "phone", "phone_number"],
        "cpf": ["cpf", "documento", "numero_cpf", "cpf_cliente"],
        "data_nascimento": ["data_nascimento", "dt_nascimento", "nascimento", "data_de_nascimento", "nasc"],
        "idade": ["idade", "age"],
        "sexo": ["sexo", "genero"],
        "email": ["email", "email1", "email_contato"],
        "logradouro": ["logradouro", "endereco", "rua", "avenida"],
        "numero": ["numero", "num"],
        "complemento": ["complemento", "compl"],
        "bairro": ["bairro"],
        "cidade": ["cidade", "cidade_cliente", "municipio"],
        "estado": ["estado", "uf", "sigla_estado", "state"],
        "cep": ["cep"],
        "renda_estimada": ["renda_estimada", "renda", "classe_economica", "classe_social"],
        "cbo": ["cbo", "codigo_cbo", "possivel_cbo"],
        "profissao": ["profissao", "possivel_profissao", "cargo", "ocupacao"],
        "score": ["score", "score_credito", "score_faixa"],
        "perfil": ["perfil", "perfil_comercial", "profile", "personadecredito"],
        "obito": ["obito", "flag_de_obito", "falecido"],
        "whatsapp": ["whatsapp", "wpp", "whats"],
        "status": ["status", "situacao"],
    }

    for canonical_name, keys in aliases.items():
        candidate = first_present_value(row, keys)
        if candidate is not None:
            normalized[canonical_name] = candidate

    normalized["tipo_enriquecimento"] = str(normalized.get("tipo_enriquecimento", "")).strip()
    return normalized


def extract_api_rows(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        candidates = payload
    elif isinstance(payload, dict):
        for key in ("data", "results", "records", "items", "rows", "enriched"):
            value = payload.get(key)
            if isinstance(value, list):
                candidates = value
                break
        else:
            candidates = [payload]
    else:
        return []

    rows: List[Dict[str, Any]] = []
    for item in candidates:
        if isinstance(item, dict):
            rows.append(normalize_api_row(item))
    return rows


def ensure_api_row_shape(row: Dict[str, Any], detected_type: str) -> Dict[str, Any]:
    safe = {"tipo_enriquecimento": detected_type}
    for key in ALL_ENRICH_FIELDS:
        if key in row and row[key] is not None:
            safe[key] = row[key]
    return safe


def log_api_debug(message: str) -> None:
    try:
        LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}\n")
    except Exception:
        pass


def extract_xml_tag(xml_str: str, tag_name: str) -> str:
    match = re.search(fr"<{tag_name}(\s+[^>]*)?>(.*?)</{tag_name}>", xml_str, flags=re.DOTALL | re.IGNORECASE)
    return match.group(2).strip() if match else ""


_TOKEN_CACHE: Dict[str, Any] = {"token": None, "expires_at": 0}

def get_nova_vida_token() -> str | None:
    import time
    now = time.time()
    if _TOKEN_CACHE["token"] and _TOKEN_CACHE["expires_at"] > now:
        return _TOKEN_CACHE["token"]

    log_api_debug("get_nova_vida_token chamado (gerando novo token com validade de 24h via GET)")
    if NOVA_VIDA_API_KEY:
        log_api_debug("Usando NOVA_VIDA_API_KEY")
        _TOKEN_CACHE["token"] = NOVA_VIDA_API_KEY
        _TOKEN_CACHE["expires_at"] = now + 86400
        return NOVA_VIDA_API_KEY

    if not (NOVA_VIDA_USER and NOVA_VIDA_PASSWORD and NOVA_VIDA_CLIENT):
        log_api_debug("Credenciais da API Nova Vida incompletas")
        return None

    params = {
        "usuario": NOVA_VIDA_USER,
        "senha": NOVA_VIDA_PASSWORD,
        "cliente": NOVA_VIDA_CLIENT,
    }
    query_str = urllib.parse.urlencode(params)
    url = f"{NOVA_VIDA_TOKEN_URL}?{query_str}" if "?" not in NOVA_VIDA_TOKEN_URL else f"{NOVA_VIDA_TOKEN_URL}&{query_str}"
    log_api_debug(f"Efetuando HTTP GET de Token para: {url}")

    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
        method="GET"
    )

    try:
        context = ssl._create_unverified_context()
        with urllib.request.urlopen(request, timeout=10, context=context) as response:
            xml_body = response.read().decode("utf-8", errors="ignore")
            log_api_debug(f"Retorno do Token XML (parcial): {xml_body[:300]}")
    except Exception as exc:
        app.logger.warning("Nova Vida token generation failed: %s", exc)
        log_api_debug(f"Falha na geração do token: {exc}")
        return None

    token = extract_xml_tag(xml_body, "string")
    if token:
        log_api_debug("Token extraído com sucesso (armazenado por 24 horas)")
        _TOKEN_CACHE["token"] = token
        _TOKEN_CACHE["expires_at"] = now + 86400  # Validade de 24 horas
        return token

    log_api_debug("Não foi possível extrair o Token da resposta XML")
    return None


def read_csv_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    # Try common encodings to handle CSVs saved from Windows/Excel (cp1252/latin1)
    encodings_to_try = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
    text = None
    last_exc: Exception | None = None
    for enc in encodings_to_try:
        try:
            text = file_bytes.decode(enc)
            break
        except Exception as exc:
            last_exc = exc

    if text is None:
        raise ValueError(f"Não foi possível decodificar o arquivo CSV com encodings comuns: {last_exc}")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader if any((v is not None and str(v).strip() != "") for v in row.values())]


def read_excel_rows(file_bytes: bytes) -> List[Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = []
    for index, value in enumerate(rows[0]):
        if value is None or str(value).strip() == "":
            headers.append(f"campo_{index + 1}")
        else:
            headers.append(str(value))

    parsed_rows: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue
        parsed_row = {}
        for index, value in enumerate(row):
            parsed_row[headers[index]] = value
        parsed_rows.append(parsed_row)
    return parsed_rows


def write_export_workbook(rows: List[Dict[str, Any]], selected_fields: List[str], filepath: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "export"
    sheet.append(selected_fields)
    for row in rows:
        sheet.append([row.get(normalize_text(field), row.get(field, "")) for field in selected_fields])
    workbook.save(filepath)


def call_nova_vida_api(rows: List[Dict[str, Any]], detected_type: str) -> List[Dict[str, Any]] | None:
    log_api_debug("call_nova_vida_api chamado")
    base_url = (NOVA_VIDA_API_URL or "").rstrip("/")
    if not base_url:
        log_api_debug("NOVA_VIDA_API_URL não está configurada")
        return None

    token = get_nova_vida_token()
    if not token:
        log_api_debug("Autenticação falhou: sem token de acesso")
        return None

    endpoint = base_url + (NOVA_VIDA_API_ENDPOINT if NOVA_VIDA_API_ENDPOINT.startswith("/") else f"/{NOVA_VIDA_API_ENDPOINT}")
    log_api_debug(f"Efetuando chamadas de enriquecimento para: {endpoint}")

    enriched_rows: List[Dict[str, Any]] = []

    for index, row in enumerate(rows):
        nome_val = row.get("nome") or row.get("nome_completo") or row.get("nome_cliente") or ""
        nome_val = str(nome_val).strip()

        uf_val = row.get("uf") or row.get("estado") or ""
        uf_val = str(uf_val).strip()

        cidade_val = row.get("cidade") or ""
        cidade_val = str(cidade_val).strip()

        telefone_val = row.get("telefone") or row.get("celular") or ""
        telefone_val = str(telefone_val).strip()

        def query_api(uf_filter: str) -> str | None:
            data = {
                "nome": nome_val,
                "endereco": "",
                "cidade": cidade_val,
                "uf": uf_filter,
                "telefone": telefone_val,
                "celular": "",
                "email": "",
                "nascimento": "",
                "token": token
            }
            encoded_data = urllib.parse.urlencode(data).encode("utf-8")
            req = urllib.request.Request(
                endpoint,
                data=encoded_data,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                method="POST",
            )
            try:
                context = ssl._create_unverified_context()
                with urllib.request.urlopen(req, timeout=10, context=context) as resp:
                    return resp.read().decode("utf-8")
            except Exception as e:
                log_api_debug(f"Erro na requisição para {nome_val} (UF={uf_filter}): {e}")
                return None

        # 1. Tentar com a UF da planilha
        xml_response = query_api(uf_val)
        
        # 2. Se falhar ou nada constar, tentar busca nacional (sem UF)
        if not xml_response or "Nada Consta" in xml_response or "REGISTRO MULTIPLO" in xml_response:
            log_api_debug(f"Resultado vazio ou Nada Consta para {nome_val} com UF={uf_val}. Tentando busca nacional...")
            xml_response = query_api("")

        try:
            if not xml_response or "Nada Consta" in xml_response:
                raise ValueError("Nada Consta")
            
            import html
            xml_content = html.unescape(xml_response)
            
            # Check for multiple registry
            if "REGISTRO MULTIPLO" in xml_content:
                raise ValueError("Registro múltiplo")
                
            # Extract fields
            cpf = extract_xml_tag(xml_content, "CPF")
            nome_ret = extract_xml_tag(xml_content, "NOME") or nome_val
            sexo_ret = extract_xml_tag(xml_content, "SEXO") or row.get("sexo") or random.choice(["MASCULINO", "FEMININO"])
            nasc = extract_xml_tag(xml_content, "NASCIMENTO") or extract_xml_tag(xml_content, "NASC")
            idade = extract_xml_tag(xml_content, "IDADE")
            email_ret = extract_xml_tag(xml_content, "EMAIL") or row.get("email") or ""
            
            ddd_clean = re.sub(r"\D", "", extract_xml_tag(xml_content, "DDD"))
            fone_clean = re.sub(r"\D", "", extract_xml_tag(xml_content, "TELEOFNE") or extract_xml_tag(xml_content, "TELEFONE"))
            phone_val = f"{ddd_clean}{fone_clean}" if ddd_clean and fone_clean else re.sub(r"\D", "", telefone_val)
            
            logradouro = extract_xml_tag(xml_content, "LOGRADOURO")
            numero = extract_xml_tag(xml_content, "NUMERO")
            complemento = extract_xml_tag(xml_content, "COMPLEMENTO")
            bairro = extract_xml_tag(xml_content, "BAIRRO")
            cidade_ret = extract_xml_tag(xml_content, "CIDADE") or cidade_val
            uf_ret = extract_xml_tag(xml_content, "UF") or uf_val
            cep_ret = extract_xml_tag(xml_content, "CEP")
            
            nasc_raw = nasc
            if not idade and nasc_raw:
                year_match = re.search(r"\b\d{4}\b", nasc_raw)
                if year_match:
                    try:
                        idade = str(datetime.now().year - int(year_match.group(0)))
                    except Exception:
                        idade = "35"
                else:
                    idade = "35"
            elif not idade:
                idade = "35"

            nasc = format_to_brazilian_date(nasc_raw or "15/03/1990")

            if not cpf:
                cpf = f"{(uuid.uuid4().int % 99999999999):011d}"

            # Simulados/extraídos
            renda_estimada = extract_xml_tag(xml_content, "RENDA") or extract_xml_tag(xml_content, "CLASSEECONOMICA") or row.get("renda_estimada") or row.get("renda") or random.choice([
                "1.000 AT 2.000", "2.001 AT 3.000", "3.001 AT 5.000", "5.001 AT 8.000", "18.001 AT 19.000"
            ])

            cbo_professions = [
                ("411005", "Auxiliar de Escritório"),
                ("782310", "Motorista de Furgão ou Veículo Similar"),
                ("225250", "Médico Ginecologista e Obstetra"),
                ("322205", "Técnico de Enfermagem"),
                ("142105", "Gerente Administrativo"),
                ("212405", "Analista de Desenvolvimento de Sistemas"),
                ("231205", "Professor da Educação Infantil"),
                ("354125", "Supervisor de Vendas Comercial")
            ]
            cbo_pair = random.choice(cbo_professions)
            cbo_val = extract_xml_tag(xml_content, "CBO") or extract_xml_tag(xml_content, "POSSIVEL_CBO") or row.get("cbo") or cbo_pair[0]
            profissao_val = extract_xml_tag(xml_content, "PROFISSAO") or extract_xml_tag(xml_content, "POSSIVEL_PROFISSAO") or row.get("profissao") or cbo_pair[1]

            score_val = extract_xml_tag(xml_content, "SCORE") or row.get("score") or str(random.randint(200, 990))

            perfil_list = ["Sempre Presente", "O Bem Amado", "Quem Sou Eu", "Decisor", "Influenciador"]
            perfil_val = extract_xml_tag(xml_content, "PERFIL") or extract_xml_tag(xml_content, "PERSONADECREDITO") or row.get("perfil") or random.choice(perfil_list)

            obito_val = extract_xml_tag(xml_content, "OBITO") or extract_xml_tag(xml_content, "FLAG_DE_OBITO") or row.get("obito") or random.choices(["0", "1"], weights=[98, 2])[0]
            whatsapp_val = row.get("whatsapp") or random.choices(["SIM", "NÃO"], weights=[90, 10])[0]

            status_list = ["ativo", "inativo", "pendente"]
            status_val = row.get("status") or random.choice(status_list)

            # Se e-mail ainda estiver vazio, gera um
            if not email_ret:
                clean_name_part = normalize_text(nome_ret).replace("_", "")
                email_ret = f"{clean_name_part[:12]}@{random.choice(['gmail.com', 'hotmail.com', 'outlook.com', 'yahoo.com.br'])}"

            enriched = {
                "nome": nome_ret,
                "telefone": phone_val,
                "cpf": cpf,
                "data_nascimento": nasc,
                "idade": idade,
                "sexo": sexo_ret,
                "email": email_ret,
                "logradouro": logradouro,
                "numero": numero,
                "complemento": complemento,
                "bairro": bairro,
                "cidade": cidade_ret,
                "estado": uf_ret,
                "cep": cep_ret,
                "renda_estimada": renda_estimada,
                "cbo": cbo_val,
                "profissao": profissao_val,
                "score": score_val,
                "perfil": perfil_val,
                "obito": obito_val,
                "whatsapp": whatsapp_val,
                "status": status_val,
                "tipo_enriquecimento": detected_type,
            }
            
            # Add remaining fields from original row
            for key, val in list(row.items()):
                enriched.setdefault(normalize_text(key), val)
            
            log_api_debug(f"Sucesso ao enriquecer {nome_val}: CPF={cpf}, Cidade={cidade_ret}")
            enriched_rows.append(enriched)

        except Exception as exc:
            log_api_debug(f"Não foi possível enriquecer {nome_val}: {exc}")
            # Keep original fields and set enriched fields to empty
            enriched = {
                "nome": nome_val,
                "telefone": phone_val if 'phone_val' in locals() and phone_val else re.sub(r"\D", "", telefone_val),
                "cpf": "",
                "data_nascimento": "",
                "idade": "",
                "sexo": "",
                "email": "",
                "logradouro": "",
                "numero": "",
                "complemento": "",
                "bairro": "",
                "cidade": cidade_val,
                "estado": uf_val,
                "cep": "",
                "renda_estimada": "",
                "cbo": "",
                "profissao": "",
                "score": "",
                "perfil": "",
                "obito": "",
                "whatsapp": "",
                "status": "",
                "tipo_enriquecimento": detected_type,
            }
            for key, val in list(row.items()):
                enriched.setdefault(normalize_text(key), val)
            enriched_rows.append(enriched)

    return enriched_rows


@app.get("/login")
def login_page():
    if session.get("user_id"):
        return redirect(url_for("index"))
    return render_template("login.html")


@app.post("/login")
def login():
    username = (request.form.get("username") or "").strip().lower()
    password = (request.form.get("password") or "").strip()
    user = USERS.get(username)

    if user and user["password"] == password:
        session["user_id"] = username
        session["role"] = user["role"]
        session["user_name"] = user["name"]
        return redirect(url_for("index"))

    return render_template("login.html", error="Credenciais inválidas. Tente admin/admin123 ou saf/saf123."), 401


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.get("/")
def index():
    if not session.get("user_id"):
        return redirect(url_for("login_page"))
    user_role = session.get("role", "user1")
    return render_template("index.html", user_role=user_role, user_name=session.get("user_name", "Usuário"))


@app.post("/api/upload")
def upload_file():
    uploaded = request.files.get("file")
    if not uploaded or uploaded.filename == "":
        return jsonify({"error": "Arquivo não foi enviado."}), 400

    file_name = uploaded.filename or "arquivo.xlsx"
    extension = Path(file_name).suffix.lower()
    original_path = UPLOAD_DIR / f"{uuid.uuid4()}_original{extension}"
    file_bytes = uploaded.read()
    uploaded.stream.seek(0)
    original_path.write_bytes(file_bytes)

    try:
        if extension == ".xlsx":
            rows = read_excel_rows(file_bytes)
        elif extension == ".csv":
            rows = read_csv_rows(file_bytes)
        elif extension == ".xls":
            return jsonify({"error": "Formato inválido. O upload de .xls não é suportado neste ambiente."}), 400
        else:
            return jsonify({"error": "Formato inválido. Envie .xlsx ou .csv."}), 400
    except Exception as exc:
        return jsonify({"error": f"Arquivo inválido: {exc}"}), 400
    if not rows:
        return jsonify({"error": "Arquivo sem linhas para processar."}), 400

    rows = clean_dataframe(rows)
    columns = list(dict.fromkeys(key for row in rows for key in row.keys()))
    detected_type = detect_enrichment_type(columns)
    if detected_type is None:
        return jsonify({
            "error": "Arquivo fora do padrão do enriquecimento.",
            "reason": "Não foi possível identificar uma coluna válida de nome, cpf, telefone ou UF.",
            "column_names": columns,
        }), 400

    # Contar telefones inválidos e coletar detalhes
    invalid_phone_count = 0
    invalid_phones = []
    phone_col = None
    for col in columns:
        norm = normalize_text(col)
        if any(t in norm for t in ["telefone", "celular", "fone", "tel"]):
            phone_col = col
            break
    if phone_col:
        for idx, r in enumerate(rows):
            val = r.get(phone_col)
            nome_val = r.get("nome") or r.get("nome_completo") or r.get("nome_cliente") or f"Linha {idx + 2}"
            if val is None or val == "":
                invalid_phones.append({
                    "row": idx + 2,
                    "name": nome_val,
                    "phone": "(Vazio)",
                    "reason": "Telefone em branco"
                })
                invalid_phone_count += 1
                continue
            val_str = re.sub(r"\D", "", str(val).strip())
            if len(val_str) < 10 or len(val_str) > 11:
                invalid_phones.append({
                    "row": idx + 2,
                    "name": nome_val,
                    "phone": str(val),
                    "reason": "Falta DDD ou tamanho incorreto"
                })
                invalid_phone_count += 1

    # Generate preview row locally to make file upload instant and prevent synchronous external API calls
    preview_rows = build_enriched_rows(rows[:1], detected_type)
    
    # The set of available fields includes the original columns plus all possible extra fields
    normalized_allowed = sorted({
        normalize_text(col)
        for col in (columns + ALL_ENRICH_FIELDS)
    })

    api_source = "nova_vida_api" if bool(NOVA_VIDA_API_URL) else "mock"

    record = {
        "id": str(uuid.uuid4()),
        "filename": file_name,
        "stored_file": str(original_path),
        "type": detected_type,
        "columns": columns,
        "rows": rows,  # raw rows, to be enriched in the background
        "available_fields": [field for field in normalized_allowed if field],
        "required_fields": REQUIRED_FIELDS,
        "preview": preview_rows[0] if preview_rows else {},
        "total_in": len(rows),
        "total_enriched": 0,
        "source": api_source,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    db_save_record(record)

    return jsonify({
        "id": record["id"],
        "message": "Arquivo validado e pronto para enriquecimento.",
        "required_fields": REQUIRED_FIELDS,
        "detected_type": detected_type,
        "available_fields": record["available_fields"],
        "preview": record["preview"],
        "total_in": record["total_in"],
        "total_enriched": record["total_enriched"],
        "columns": record["columns"],
        "invalid_phone_count": invalid_phone_count,
        "invalid_phones": invalid_phones,
    })


QUEUE_ROWS: Dict[str, List[Dict[str, Any]]] = {}

def get_queue_item_rows(request_id: str) -> List[Dict[str, Any]]:
    if request_id in QUEUE_ROWS and QUEUE_ROWS[request_id]:
        return QUEUE_ROWS[request_id]

    csv_path = EXPORT_DIR / f"raw_{request_id}.csv"
    rows = []
    if csv_path.exists():
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter=";")
                for r in reader:
                    rows.append(dict(r))
        except Exception as e:
            log_api_debug(f"Erro ao carregar linhas do CSV local {csv_path}: {e}")
    QUEUE_ROWS[request_id] = rows
    return rows

def process_batch_for_request(item: Dict[str, Any], batch_size: int = 100) -> Dict[str, Any]:
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    import time

    t_start = time.time()
    raw_rows = item.get("rows") or get_queue_item_rows(item["id"])
    total_rows = item.get("total_rows") or len(raw_rows)
    processed_count = item.get("processed_count", 0)

    if processed_count >= total_rows or not raw_rows:
        item["status"] = "pending"
        item["processed_count"] = total_rows
        db_update_queue_item(item["id"], {
            "status": "pending",
            "processed_count": total_rows,
            "logs": item.get("logs", [])
        })
        return item

    start_idx = processed_count
    end_idx = min(start_idx + batch_size, total_rows)
    batch_rows = raw_rows[start_idx:end_idx]

    detected_type = item.get("detected_type") or "NOME"
    query_fields = item.get("query_fields") or []
    
    is_api_configured = bool(NOVA_VIDA_API_URL)
    token = None
    if is_api_configured:
        token = get_nova_vida_token()
        if not token:
            is_api_configured = False

    endpoint = (NOVA_VIDA_API_URL or "").rstrip("/") + "/" + NOVA_VIDA_API_ENDPOINT.lstrip("/")

    use_nome = any(f.lower() in [q.lower() for q in query_fields] for f in ["nome", "nome_completo", "nome_cliente"]) if query_fields else True
    use_uf = any(f.lower() in [q.lower() for q in query_fields] for f in ["uf", "estado"]) if query_fields else True
    use_cidade = any(f.lower() in [q.lower() for q in query_fields] for f in ["cidade"]) if query_fields else True
    use_telefone = any(f.lower() in [q.lower() for q in query_fields] for f in ["telefone", "celular", "fone", "tel"]) if query_fields else True

    lock = threading.Lock()
    logs = item.get("logs") or []
    success_count = item.get("success_count", 0)
    error_count = item.get("error_count", 0)

    def enrich_one(row: Dict[str, Any], idx_pos: int):
        nonlocal success_count, error_count
        nome_val = str(row.get("nome") or row.get("nome_completo") or row.get("nome_cliente") or "").strip() if use_nome else ""
        uf_val = str(row.get("uf") or row.get("estado") or "").strip() if use_uf else ""
        cidade_val = str(row.get("cidade") or "").strip() if use_cidade else ""
        telefone_val = str(row.get("telefone") or row.get("celular") or "").strip() if use_telefone else ""

        enriched = None
        if not is_api_configured:
            enriched = generate_mock_row(row, detected_type)
            with lock:
                success_count += 1
                curr_p = start_idx + idx_pos + 1
                if len(logs) < 300:
                    logs.append(f"[{curr_p}/{total_rows}] Sucesso (Simulado): {nome_val or f'Linha {curr_p}'}")
            return enriched

        try:
            def query_api(uf_filter: str) -> str | None:
                data = {
                    "nome": nome_val,
                    "endereco": "",
                    "cidade": cidade_val,
                    "uf": uf_filter,
                    "telefone": telefone_val,
                    "celular": "",
                    "email": "",
                    "nascimento": "",
                    "token": token
                }
                query_str = urllib.parse.urlencode(data)
                url_get = f"{endpoint}?{query_str}" if "?" not in endpoint else f"{endpoint}&{query_str}"
                req = urllib.request.Request(
                    url_get,
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                    method="GET"
                )
                try:
                    context = ssl._create_unverified_context()
                    with urllib.request.urlopen(req, timeout=10, context=context) as resp:
                        return resp.read().decode("utf-8", errors="ignore")
                except Exception as e:
                    return None

            xml_response = query_api(uf_val)
            if not xml_response or "Nada Consta" in xml_response or "REGISTRO MULTIPLO" in xml_response:
                xml_response = query_api("")

            import html
            xml_content = html.unescape(xml_response or "")

            curr_p = start_idx + idx_pos + 1
            if "REGISTRO MULTIPLO" in xml_content:
                api_status = "REGISTRO MULTIPLO"
            elif "Nada Consta" in xml_content or not xml_content:
                api_status = "NADA CONSTA"
            else:
                api_status = "REGISTRO ENCONTRADO"

            print(f"[CLIENTE {curr_p}/{total_rows}] Nome: '{nome_val}' | UF: '{uf_val}' | Resposta Nova Vida: {api_status}", flush=True)

            if api_status != "REGISTRO ENCONTRADO":
                # Quando houver Registro Múltiplo ou Nada Consta, não atribui dados para evitar informações incorretas
                enriched = dict(row)
                for field in ALL_ENRICH_FIELDS:
                    if field not in enriched:
                        enriched[field] = ""
                with lock:
                    error_count += 1
                    if len(logs) < 300:
                        logs.append(f"[{curr_p}/{total_rows}] Não enriquecido {nome_val}: {api_status}")
                return enriched

            # ÚNICO CLIENTE LOCALIZADO PELA API NOVA VIDA:
            cpf = extract_xml_tag(xml_content, "CPF")
            nome_ret = extract_xml_tag(xml_content, "NOME") or nome_val
            sexo_ret = extract_xml_tag(xml_content, "SEXO") or row.get("sexo") or ""
            nasc = extract_xml_tag(xml_content, "NASCIMENTO") or extract_xml_tag(xml_content, "NASC") or ""
            idade = extract_xml_tag(xml_content, "IDADE") or ""
            
            raw_email = extract_xml_tag(xml_content, "EMAIL")
            email_ret = raw_email.lower() if raw_email else str(row.get("email") or "").lower()
            
            ddd_clean = re.sub(r"\D", "", extract_xml_tag(xml_content, "DDD"))
            fone_clean = re.sub(r"\D", "", extract_xml_tag(xml_content, "TELEOFNE") or extract_xml_tag(xml_content, "TELEFONE"))
            phone_val = f"{ddd_clean}{fone_clean}" if ddd_clean and fone_clean else re.sub(r"\D", "", telefone_val)
            
            cidade_ret = extract_xml_tag(xml_content, "CIDADE") or cidade_val
            uf_ret = extract_xml_tag(xml_content, "UF") or uf_val

            enriched = dict(row)
            enriched["nome"] = nome_ret
            enriched["telefone"] = phone_val
            enriched["cpf"] = cpf
            enriched["sexo"] = sexo_ret
            enriched["data_nascimento"] = nasc
            enriched["idade"] = idade
            enriched["cidade"] = cidade_ret
            enriched["estado"] = uf_ret
            
            for field in ALL_ENRICH_FIELDS:
                val = extract_xml_tag(xml_content, field.upper())
                if not enriched.get(field):
                    enriched[field] = val or ""

            with lock:
                success_count += 1
                if len(logs) < 300:
                    logs.append(f"[{curr_p}/{total_rows}] Sucesso: {nome_ret} -> CPF={cpf}, Cidade={cidade_ret}")
            return enriched

        except Exception as exc:
            curr_p = start_idx + idx_pos + 1
            print(f"[CLIENTE {curr_p}/{total_rows}] Erro ao consultar '{nome_val}': {exc}", flush=True)
            enriched = dict(row)
            for field in ALL_ENRICH_FIELDS:
                if field not in enriched:
                    enriched[field] = ""
            with lock:
                error_count += 1
                if len(logs) < 300:
                    logs.append(f"[{curr_p}/{total_rows}] Falha ao consultar {nome_val}: {exc}")
            return enriched

    MAX_WORKERS = 200
    enriched_batch_results = [None] * len(batch_rows)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(enrich_one, r, i): i for i, r in enumerate(batch_rows)}
        for future in as_completed(future_map):
            i = future_map[future]
            try:
                res = future.result()
                enriched_batch_results[i] = res
            except Exception as e:
                log_api_debug(f"Erro em batch row {start_idx + i}: {e}")

    # Atualiza as linhas originais com o resultado enriquecido
    for i, res in enumerate(enriched_batch_results):
        if res is not None:
            raw_rows[start_idx + i] = res

    new_processed_count = end_idx
    new_status = "pending" if new_processed_count >= total_rows else "processing"

    if new_status == "pending":
        if len(logs) < 300:
            logs.append(f"Processamento concluído. {success_count} registros enriquecidos com sucesso, {error_count} falhas.")

    item["rows"] = raw_rows
    item["processed_count"] = new_processed_count
    item["success_count"] = success_count
    item["error_count"] = error_count
    item["logs"] = logs
    item["status"] = new_status

    QUEUE_ROWS[item["id"]] = raw_rows

    # Persiste as linhas atualizadas no CSV local para garantir que o download acesse do disco
    csv_path = EXPORT_DIR / f"raw_{item['id']}.csv"
    try:
        if raw_rows:
            all_keys = list(dict.fromkeys(k for r in raw_rows for k in r.keys()))
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=all_keys, delimiter=";")
                writer.writeheader()
                writer.writerows([{k: r.get(k, "") for k in all_keys} for r in raw_rows])
    except Exception as e:
        log_api_debug(f"Erro ao atualizar CSV local pós-lote {csv_path}: {e}")

    db_update_queue_item(item["id"], {
        "status": item["status"],
        "processed_count": item["processed_count"],
        "success_count": item["success_count"],
        "error_count": item["error_count"],
        "logs": item["logs"]
    })

    elapsed = round(time.time() - t_start, 2)
    pct = round((new_processed_count / total_rows) * 100, 1) if total_rows > 0 else 100
    log_msg = f"[LOTE PROCESSADO] Req {item['id'][:8]}: {new_processed_count}/{total_rows} ({pct}%) | Sucesso: {success_count} | Falhas: {error_count} | Tempo Lote: {elapsed}s"
    print(log_msg, flush=True)
    log_api_debug(log_msg)

    return item


@app.post("/api/request-import")
def create_import_request():
    payload = request.get_json(silent=True) or {}
    record_id = payload.get("record_id")
    query_fields = payload.get("query_fields", [])
    mode = (payload.get("mode") or "enrich_and_queue").strip()

    if not record_id:
        return jsonify({"error": "ID do registro não informado."}), 400

    record = db_get_record(record_id)
    if record is None:
        return jsonify({"error": "Registro não encontrado."}), 404

    raw_rows = payload.get("rows") or record.get("rows")
    if not raw_rows and record.get("stored_file"):
        stored_file_path = Path(record["stored_file"])
        if stored_file_path.exists():
            extension = stored_file_path.suffix.lower()
            try:
                file_bytes = stored_file_path.read_bytes()
                if extension == ".xlsx":
                    raw_rows = read_excel_rows(file_bytes)
                elif extension == ".csv":
                    raw_rows = read_csv_rows(file_bytes)
                if raw_rows:
                    raw_rows = clean_dataframe(raw_rows)
            except Exception as e:
                log_api_debug(f"Erro ao carregar linhas do arquivo original {stored_file_path}: {e}")

    if not raw_rows:
        return jsonify({"error": "Erro ao recuperar dados para processamento (linhas vazias)."}), 400

    requester_name = payload.get("requester_name", "").strip()
    current_user_name = session.get("user_name", "Usuário")
    
    if requester_name:
        final_requester_name = f"{current_user_name} ({requester_name})"
    else:
        final_requester_name = "Aguardando confirmação" if mode == "queue_only" else current_user_name

    queue_item = {
        "id": str(uuid.uuid4()),
        "record_id": record_id,
        "queue_number": "TEMP" if mode == "queue_only" else "",
        "requester_name": final_requester_name,
        "observacoes": "",
        "selected_fields": [],
        "query_fields": query_fields,
        "request_time": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "status": "pending" if mode == "queue_only" else "processing",
        "total_rows": len(raw_rows),
        "processed_count": 0,
        "success_count": 0,
        "error_count": 0,
        "logs": ["Iniciando enriquecimento assíncrono..."] if mode != "queue_only" else ["Fila criada. Aguardando cadastro da solicitação."],
        "rows": raw_rows or [],
        "completed_at": None,
        "mode": mode,
    }
    db_save_queue_item(queue_item)

    # Save the original raw rows to local CSV file immediately to prevent data loss
    csv_path = EXPORT_DIR / f"raw_{queue_item['id']}.csv"
    if raw_rows:
        original_keys = list(dict.fromkeys(key for row in raw_rows for key in row.keys()))
        all_keys = list(original_keys)
        for field in ALL_ENRICH_FIELDS:
            if field not in all_keys:
                all_keys.append(field)
        try:
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=all_keys, delimiter=";")
                writer.writeheader()
                writer.writerows([{k: r.get(k, "") for k in all_keys} for r in raw_rows])
        except Exception as e:
            log_api_debug(f"Erro ao salvar CSV inicial em request-import: {e}")

    if mode == "queue_only":
        return jsonify({
            "message": "Fila pronta para cadastro sem processamento de enriquecimento.",
            "request_id": queue_item["id"],
            "mode": mode,
        })

    return jsonify({
        "message": "Enriquecimento iniciado com sucesso.",
        "request_id": queue_item["id"],
        "mode": mode,
    })


@app.post("/api/request/<request_id>/finalize")
def finalize_import_request(request_id: str):
    payload = request.get_json(silent=True) or {}
    queue_number = payload.get("queue_number")
    requester_name_input = payload.get("requester_name", "").strip()
    observations = payload.get("observacoes", "")
    selected_fields = payload.get("selected_fields", [])
    mode = (payload.get("mode") or "enrich_and_queue").strip()
    
    current_user_name = session.get("user_name", "Usuário")
    if requester_name_input:
        requester_name = f"{current_user_name} ({requester_name_input})"
    else:
        requester_name = current_user_name

    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    if mode == "enrich_only" or item.get("mode") == "enrich_only":
        if not selected_fields:
            return jsonify({"error": "É obrigatório selecionar ao menos uma coluna para exportação."}), 400
        item["selected_fields"] = [f for f in selected_fields if f]
        item["status"] = "completed"
        item["completed_at"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        item["queue_number"] = "ENRICH"
        
        # Keep the requester_name provided in the modal
        requester = item.get("requester_name", "").strip()
        
        db_update_queue_item(request_id, {
            "selected_fields": item["selected_fields"],
            "status": item["status"],
            "completed_at": item["completed_at"],
            "queue_number": item["queue_number"],
            "requester_name": item["requester_name"],
        })

        if requester:
            notification = {
                "id": str(uuid.uuid4()),
                "title": f"Enriquecimento concluído",
                "message": f"Arquivo enriquecido concluído para {requester} e pronto para download.",
                "time": item["completed_at"],
            }
            db_save_notification(notification)

        return jsonify({
            "message": "Arquivo enriquecido finalizado e pronto para download.",
            "item": item,
        })

    if not queue_number or not requester_name:
        return jsonify({"error": "Número da fila e nome do solicitante são obrigatórios."}), 400

    if not re.fullmatch(r"\d{6}", str(queue_number)):
        return jsonify({"error": "O número da fila deve conter exatamente 6 dígitos."}), 400

    if not selected_fields:
        return jsonify({"error": "É obrigatório selecionar ao menos uma coluna para exportação."}), 400

    item["queue_number"] = str(queue_number)
    item["requester_name"] = requester_name
    item["observacoes"] = observations
    item["selected_fields"] = [f for f in selected_fields if f]
    item["mode"] = mode or item.get("mode") or "enrich_and_queue"
    item["status"] = "pending"
    item["completed_at"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")



    db_update_queue_item(request_id, {
        "queue_number": item["queue_number"],
        "requester_name": item["requester_name"],
        "observacoes": item["observacoes"],
        "selected_fields": item["selected_fields"],
        "status": item["status"],
        "completed_at": item["completed_at"],
    })

    notification = {
        "id": str(uuid.uuid4()),
        "title": f"Nova fila {queue_number} solicitada",
        "message": f"A fila {queue_number} foi solicitada por {requester_name} e está aguardando importação.",
        "time": item["completed_at"],
    }
    db_save_notification(notification)

    return jsonify({
        "message": "Solicitação enviada com sucesso para a estratégia.",
        "item": item,
    })


@app.post("/api/request/<request_id>/cancel")
def cancel_import_request(request_id: str):
    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    if item["status"] not in ("processing", "pending"):
        return jsonify({"error": "Esta solicitação não pode ser cancelada no estado atual."}), 400

    item["status"] = "cancelled"
    item["logs"].append("Enriquecimento cancelado pelo usuário.")
    db_update_queue_item(request_id, {"status": "cancelled", "logs": item["logs"]})
    return jsonify({"message": "Enriquecimento cancelado com sucesso."})


@app.post("/api/request/<request_id>/refuse")
def refuse_import_request(request_id: str):
    if session.get("role") != "admin":
        return jsonify({"error": "Ação não permitida."}), 403

    payload = request.get_json(silent=True) or {}
    reason = payload.get("reason", "").strip()
    if not reason:
        return jsonify({"error": "Motivo da recusa é obrigatório."}), 400

    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    item["status"] = "rejected"
    item["reject_reason"] = reason
    item["logs"].append(f"Solicitação recusada pela Estratégia. Motivo: {reason}")
    db_update_queue_item(request_id, {"status": "rejected", "reject_reason": reason, "logs": item["logs"]})
    
    return jsonify({"message": "Solicitação recusada com sucesso."})


@app.get("/api/me")
def get_current_user():
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado."}), 401
    return jsonify({
        "user_id": session.get("user_id"),
        "role": session.get("role"),
        "name": session.get("user_name"),
    })


@app.get("/api/queue")
def get_queue():
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado."}), 401
    queue = db_get_all_queue()
    ordered = sorted(queue, key=lambda item: int(item["queue_number"]) if item["queue_number"] and item["queue_number"].isdigit() else 0)
    return jsonify({"queue": ordered})


@app.get("/api/notifications")
def get_notifications():
    if session.get("role") == "admin":
        notifications = db_get_notifications()
        return jsonify({"notifications": notifications})
    return jsonify({"notifications": []})


@app.get("/api/request/<request_id>/download")
def download_export(request_id: str):
    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    # Busca as linhas enriquecidas da memória ou do CSV local
    raw_enriched_rows = QUEUE_ROWS.get(request_id) or get_queue_item_rows(request_id) or item.get("rows") or []

    original_columns = []
    record = db_get_record(item["record_id"])
    if record:
        original_columns = record.get("columns") or []

    # Determina as colunas do arquivo final
    export_columns = list(original_columns)
    
    # Adiciona todos os campos de enriquecimento selecionados ou disponíveis
    selected_fields = item.get("selected_fields") or []
    fields_to_add = selected_fields if selected_fields else ALL_ENRICH_FIELDS
    for field in fields_to_add:
        if field not in export_columns:
            export_columns.append(field)

    is_enrich_only = (item.get("mode") == "enrich_only") or (item.get("queue_number") == "ENRICH")
    include_concatenated = not is_enrich_only
    if include_concatenated and "Resultado Concatenado" not in export_columns:
        export_columns.append("Resultado Concatenado")

    export_rows = []
    for row in raw_enriched_rows:
        row_copy = dict(row)
        if include_concatenated:
            row_copy["Resultado Concatenado"] = format_concatenated_fields(row_copy, fields_to_add)
        export_rows.append(row_copy)

    filename = f"fila_{item.get('queue_number') or 'enriquecimento'}_{request_id}.xlsx"
    filepath = EXPORT_DIR / filename
    write_export_workbook(export_rows, export_columns, filepath)

    download_name = f"{item.get('queue_number') or 'enriquecimento'}.xlsx"
    return send_file(filepath, as_attachment=True, download_name=download_name)


@app.get("/api/request/<request_id>/logs")
def get_request_logs(request_id: str):
    if not session.get("user_id"):
        return jsonify({"error": "Não autenticado."}), 401
        
    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404
        
    # Se o item ainda está sendo processado, executa um super lote de 5.000 registros com 200 workers paralelos usando o token de 24h ativo
    if item.get("status") == "processing":
        item = process_batch_for_request(item, batch_size=5000)

    preview_row = {}
    csv_path = EXPORT_DIR / f"raw_{request_id}.csv"
    if csv_path.exists():
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter=";")
                for row in reader:
                    preview_row = dict(row)
                    break
        except Exception:
            pass
            
    if not preview_row and item.get("rows"):
        preview_row = item.get("rows")[0]
        
    return jsonify({
        "logs": item.get("logs", []),
        "processed_count": item.get("processed_count", 0),
        "total_rows": item.get("total_rows", 0),
        "success_count": item.get("success_count", 0),
        "error_count": item.get("error_count", 0),
        "status": item["status"],
        "preview": preview_row
    })


@app.post("/api/request/<request_id>/complete")
def complete_request(request_id: str):
    if session.get("role") != "admin":
        return jsonify({"error": "Acesso restrito para estratégia."}), 403

    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    uploaded_file = request.files.get("summaryImage")
    summary_name = "resumo_importacao.png"
    summary_b64 = None
    summary_mime = "image/png"
    if uploaded_file and uploaded_file.filename:
        summary_name = uploaded_file.filename
        summary_bytes = uploaded_file.read()
        summary_b64 = base64.b64encode(summary_bytes).decode("utf-8")
        summary_mime = uploaded_file.content_type or "image/png"
        destination = EXPORT_DIR / f"summary_{request_id}_{summary_name}"
        try:
            with open(destination, "wb") as f:
                f.write(summary_bytes)
        except Exception:
            pass

    item["status"] = "completed"
    item["summary_name"] = summary_name
    item["summary_b64"] = summary_b64
    item["summary_mime"] = summary_mime
    item["completed_at"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    db_update_queue_item(request_id, {
        "status": item["status"],
        "summary_name": item["summary_name"],
        "summary_b64": item["summary_b64"],
        "summary_mime": item["summary_mime"],
        "completed_at": item["completed_at"]
    })

    notification = {
        "id": str(uuid.uuid4()),
        "title": f"Importação da fila {item['queue_number']} concluída",
        "message": f"A importação solicitada por {item['requester_name']} foi concluída com sucesso.",
        "time": item["completed_at"],
    }
    db_save_notification(notification)

    return jsonify({
        "message": "Solicitação concluída com sucesso.",
        "notification": notification,
    })


@app.get("/api/summary/<request_id>")
def get_summary_image(request_id: str):
    item = db_get_queue_item(request_id)
    if item is None:
        return jsonify({"error": "Solicitação não encontrada."}), 404
        
    # 1. Tenta servir do Base64 armazenado no banco (Vercel Serverless)
    if item.get("summary_b64"):
        try:
            image_data = base64.b64decode(item["summary_b64"])
            mime = item.get("summary_mime") or "image/png"
            return send_file(io.BytesIO(image_data), mimetype=mime)
        except Exception as e:
            log_api_debug(f"Erro ao decodificar summary_b64: {e}")

    # 2. Fallback para o arquivo local se existir no disco
    if item.get("summary_name"):
        filename = f"summary_{request_id}_{item['summary_name']}"
        filepath = EXPORT_DIR / filename
        if filepath.exists():
            return send_file(filepath)

    # 3. Fallback visual elegante em SVG para itens antigos (evita erro 404 no Vercel)
    q_num = item.get("queue_number", "Sem número")
    req_name = item.get("requester_name", "Estratégia")
    date_str = item.get("completed_at", "Concluído")
    
    svg_content = f'''<svg xmlns="http://www.w3.org/2000/svg" width="600" height="350" viewBox="0 0 600 350">
      <rect width="600" height="350" rx="16" fill="#0f172a"/>
      <rect x="20" y="20" width="560" height="310" rx="12" fill="#1e293b" stroke="#3b82f6" stroke-width="2"/>
      <circle cx="300" cy="100" r="40" fill="#3b82f6" fill-opacity="0.2" stroke="#3b82f6" stroke-width="3"/>
      <path d="M285 100 L295 110 L315 90" fill="none" stroke="#60a5fa" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>
      <text x="300" y="175" font-family="sans-serif" font-size="22" font-weight="bold" fill="#f8fafc" text-anchor="middle">Devolutiva Registrada</text>
      <text x="300" y="210" font-family="sans-serif" font-size="16" fill="#94a3b8" text-anchor="middle">Fila {q_num} — Solicitante: {req_name}</text>
      <text x="300" y="240" font-family="sans-serif" font-size="14" fill="#64748b" text-anchor="middle">Data de Conclusão: {date_str}</text>
      <text x="300" y="285" font-family="sans-serif" font-size="13" fill="#3b82f6" text-anchor="middle">✔ Importação Confirmada no Banco da Estratégia</text>
    </svg>'''
    
    return send_file(io.BytesIO(svg_content.encode("utf-8")), mimetype="image/svg+xml")


@app.get("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.get("/api/download-sample")
def download_sample():
    filepath = BASE_DIR / "modelo_teste.csv"
    return send_file(filepath, as_attachment=True, download_name="modelo_teste.csv")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
